"""
Export the full 34-week training plan to Excel.
Run locally: python scripts/export_plan.py
Run on VM:   python scripts/export_plan.py  (saves to data/training_plan.xlsx)
"""
import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agent.db import get_connection
from agent import config

# ── Colours per workout type ───────────────────────────────────────────────────
COLOURS = {
    "easy":      "D9EAD3",   # soft green
    "long":      "B6D7A8",   # medium green
    "tempo":     "FFE599",   # yellow
    "intervals": "F4CCCC",   # red-ish
    "race":      "EA9999",   # salmon
    "rest":      "F3F3F3",   # grey (unused rows)
}

PHASE_LABELS = {
    range(1, 11):  "Phase 1 — Base",
    range(11, 23): "Phase 2 — Build",
    range(23, 31): "Phase 3 — Peak",
    range(31, 35): "Phase 4 — Taper",
}

DAY_ORDER = {"monday": 1, "wednesday": 2, "saturday": 3}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def phase_for(week: int) -> str:
    for r, label in PHASE_LABELS.items():
        if week in r:
            return label
    return ""


def export(output_path: str):
    cfg = config.load()
    start_date   = datetime.date.fromisoformat(cfg["training"]["plan_start_date"])
    marathon_date = datetime.date.fromisoformat(cfg["training"]["marathon_date"])

    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM training_plan ORDER BY week_number, id"
        ).fetchall()

    workouts = [dict(r) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Plan"

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Marathon Training Plan — {cfg['user']['name']}  |  Race: {marathon_date.strftime('%d %b %Y')}"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="274E13")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header ─────────────────────────────────────────────────────────────────
    headers = ["Week", "Phase", "Date (Mon)", "Day", "Type", "Distance (km)", "Pace Target (/km)", "Description", "✓ Done"]
    ws.append(headers)
    header_row = ws[2]
    for cell in header_row:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="38761D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[2].height = 20

    # ── Data ───────────────────────────────────────────────────────────────────
    current_week = None
    row_num = 3

    # group by week
    from itertools import groupby
    for week_num, week_workouts in groupby(workouts, key=lambda x: x["week_number"]):
        week_workouts = sorted(week_workouts, key=lambda x: DAY_ORDER.get(x["day_of_week"], 9))
        week_start = start_date + datetime.timedelta(weeks=week_num - 1)
        phase      = phase_for(week_num)
        colour     = None

        for wo in week_workouts:
            colour = COLOURS.get(wo["workout_type"], "FFFFFF")
            fill   = PatternFill("solid", fgColor=colour)

            row = [
                week_num,
                phase,
                week_start.strftime("%d/%m/%Y"),
                wo["day_of_week"].capitalize(),
                wo["workout_type"].capitalize(),
                wo["distance_km"],
                wo["pace_target"],
                wo["description"],
                "",   # done checkbox
            ]
            ws.append(row)
            for cell in ws[row_num]:
                cell.fill      = fill
                cell.border    = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws[f"A{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"F{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"G{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"I{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        # thin separator between weeks
        ws.row_dimensions[row_num - 1].height = 18

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = [7, 22, 14, 13, 13, 16, 20, 55, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Legend ─────────────────────────────────────────────────────────────────
    legend_row = row_num + 1
    ws[f"A{legend_row}"].value = "Legend:"
    ws[f"A{legend_row}"].font  = Font(bold=True)
    for i, (wtype, colour) in enumerate(COLOURS.items()):
        if wtype == "rest":
            continue
        col = i + 2
        cell = ws.cell(row=legend_row, column=col, value=wtype.capitalize())
        cell.fill      = PatternFill("solid", fgColor=colour)
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center")

    # ── Freeze top rows ────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "data", "training_plan.xlsx"
    )
    export(out)
